import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

WHITE       = 'FFFFFFFF'
LIGHT_GRAY  = 'FFF8FAFC'
RED_SOFT    = 'FFFEE2E2'
BLUE_SOFT   = 'FFdbeafe'
GREEN_SOFT  = 'FFdcfce7'
SLATE_SOFT  = 'FFF1F5F9'

SECTION_COLORS = {
    'Bugs / Open Issues':                   ('FFFEE2E2', 'FF991b1b'),
    'Settings - Teams & Users':             ('FFe0f2fe', 'FF0369a1'),
    'Asset Catalog':                        ('FFfce7f3', 'FF9d174d'),
    'Asset Profile':                        ('FFfdf4ff', 'FF7e22ce'),
    'Data Quality':                         ('FFfef3c7', 'FFb45309'),
    'Tag Management':                       ('FFd1fae5', 'FF065f46'),
    'Settings - Connections':               ('FFe0e7ff', 'FF3730a3'),
    'Access Control':                       ('FFffe4e6', 'FF9f1239'),
    'Settings - SSO':                       ('FFf0fdf4', 'FF166534'),
    'Connectors - Metadata Ingestion':      ('FFfff7ed', 'FFc2410c'),
}

def thin_border():
    s = Side(style='thin', color='FFe2e8f0')
    return Border(left=s, right=s, top=s, bottom=s)

def cell(ws, row, col, value, bold=False, bg=None, fg='FF1E293B',
         size=10, wrap=False, align='left', italic=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, color=fg, italic=italic, name='Calibri')
    if bg:
        c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = thin_border()
    return c

priority_colors = {'High': 'FFEF4444', 'Medium': 'FFF97316', 'Low': 'FF94A3B8'}
status_bg = {'New': SLATE_SOFT, 'Open': RED_SOFT, 'In Progress': BLUE_SOFT, 'Done': GREEN_SOFT}

# ════════════════════════════════════════
# SHEET 1 — ACTIVE STORIES
# ════════════════════════════════════════
ws = wb.active
ws.title = 'Active Stories'
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'A3'

for i, w in enumerate([10, 30, 64, 11, 12, 22, 28], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[1].height = 36
ws.row_dimensions[2].height = 20

ws.merge_cells('A1:G1')
t = ws['A1']
t.value = 'EDG  –  May 2026 Sprint  |  Active User Stories'
t.font = Font(bold=True, size=14, color=WHITE, name='Calibri')
t.fill = PatternFill('solid', fgColor='FF0F172A')
t.alignment = Alignment(horizontal='center', vertical='center')

for i, h in enumerate(['Jira ID', 'Section', 'Story Title', 'Priority', 'Status', 'Assignee', 'Notes'], 1):
    cell(ws, 2, i, h, bold=True, bg='FF334155', fg=WHITE, size=10, align='center')

sections = [
    ('Bugs / Open Issues', [
        ('EDG-531', 'Non-admin user is able to view Invite User button on Users page', 'High', 'Open', 'rambabu kondapaturi'),
        ('EDG-534', 'Default role is not pre-selected in Add User modal', 'High', 'Open', 'rambabu kondapaturi'),
    ]),
    ('Settings - Teams & Users', [
        ('EDG-53',  'Search and Filter the User Directory', 'High', 'New', 'Unassigned'),
        ('EDG-665', 'Add User modal redesign - Work Email, Default Role, Team, Password fields', 'High', 'New', 'Unassigned'),
        ('EDG-666', 'RoleChipPicker - multi-role support with star primary role and popover actions', 'High', 'New', 'Unassigned'),
        ('EDG-667', 'Users table pagination - Showing X-Y of N members with page buttons', 'High', 'New', 'Unassigned'),
        ('EDG-668', 'Login page - remove role selector; derive role from user account automatically', 'High', 'New', 'Unassigned'),
        ('EDG-672', 'Activate/Deactivate User and Remove User from Platform', 'High', 'New', 'Unassigned'),
        ('EDG-673', 'Team Detail Panel - View Members, Edit Team, Add/Remove Members, Delete Team', 'High', 'New', 'Unassigned'),
        ('EDG-674', 'Teams tab search & pagination + inline team assignment in Users table', 'High', 'New', 'Unassigned'),
    ]),
    ('Asset Catalog', [
        ('EDG-69',  'Browse Asset Catalog List View', 'High', 'New', 'Unassigned'),
        ('EDG-70',  'Filter Assets by Connector, Domain, Type, Certification, Tier, Owner, Tags, Glossary', 'High', 'New', 'Unassigned'),
        ('EDG-71',  'Search Assets by Name, Description, Tags, and Path', 'High', 'New', 'Unassigned'),
        ('EDG-84',  'Sort Assets in Catalog', 'Medium', 'New', 'Unassigned'),
        ('EDG-546', 'Toggle Asset Catalog view between table and grid layout', 'Medium', 'New', 'Unassigned'),
        ('EDG-651', 'Navigate into Databases and Schemas via Asset Drilldown', 'High', 'New', 'Unassigned'),
    ]),
    ('Asset Profile', [
        ('EDG-73',  'Navigation Shell, Tab Bar & Sidebar', 'High', 'New', 'Unassigned'),
        ('EDG-74',  'Edit Asset Owners via Multi-Select in Asset Profile Sidebar', 'High', 'New', 'Unassigned'),
        ('EDG-75',  'Add and Remove Asset Tags in Asset Profile Sidebar', 'High', 'New', 'Unassigned'),
        ('EDG-85',  'View Overview Tab', 'High', 'New', 'Unassigned'),
        ('EDG-86',  'View & Manage Schema Tab (Column List + Column Detail Panel)', 'High', 'New', 'Unassigned'),
        ('EDG-88',  'View Quality Tab', 'High', 'New', 'Unassigned'),
        ('EDG-90',  'Edit Asset Stewards via Multi-Select in Asset Profile Sidebar', 'High', 'New', 'Unassigned'),
        ('EDG-551', 'Edit asset tier assignment from the asset profile', 'Medium', 'New', 'Unassigned'),
        ('EDG-552', 'Assign and remove glossary terms from the asset profile', 'High', 'New', 'Unassigned'),
        ('EDG-553', 'View asset usage statistics tab in the asset profile', 'Medium', 'New', 'Unassigned'),
    ]),
    ('Data Quality', [
        ('EDG-554', 'View asset-level DQ score, summary, and trend on the asset profile', 'High', 'New', 'Unassigned'),
        ('EDG-555', 'DQ Tab - Select Test Cases for a Table', 'High', 'New', 'Unassigned'),
        ('EDG-556', 'Column Panel - Select Test Cases for a Column', 'High', 'New', 'Unassigned'),
        ('EDG-557', 'Add a Test Case (All Supported Test Types)', 'High', 'New', 'Unassigned'),
        ('EDG-558', 'Run individual, selected, or all test cases on an asset', 'High', 'New', 'Unassigned'),
        ('EDG-559', 'View test case result details and failure reason', 'High', 'New', 'Unassigned'),
        ('EDG-560', 'Edit or delete an existing test case', 'Medium', 'New', 'Unassigned'),
        ('EDG-562', 'Filter test cases by DQ dimension and execution status', 'Medium', 'New', 'Unassigned'),
        ('EDG-563', 'Acknowledge and resolve a DQ incident from the test case view', 'High', 'New', 'Unassigned'),
        ('EDG-670', 'Incident column with View badge on Asset Quality Tab test case table', 'Medium', 'New', 'Unassigned'),
    ]),
    ('Tag Management', [
        ('EDG-564', 'Browse and filter tags in the Tag Management module', 'High', 'New', 'Unassigned'),
        ('EDG-565', 'Create a new tag definition via the slide-in drawer', 'High', 'New', 'Unassigned'),
        ('EDG-566', 'Edit an existing tag definition inline', 'Medium', 'New', 'Unassigned'),
        ('EDG-567', 'Delete a tag definition from Tag Management', 'Medium', 'New', 'Unassigned'),
        ('EDG-568', 'Apply a tag to an asset from the asset profile', 'High', 'New', 'Unassigned'),
        ('EDG-569', 'Remove a tag from an asset on the asset profile', 'High', 'New', 'Unassigned'),
        ('EDG-570', 'View assets linked to a tag via the Linked Assets tab', 'Medium', 'New', 'Unassigned'),
        ('EDG-571', 'Create a new tag category via the slide-in drawer', 'Medium', 'New', 'Unassigned'),
    ]),
    ('Settings - Connections', [
        ('EDG-573', '[Connections Overview] View all connections with live health status, filters, and search', 'High', 'New', 'Unassigned'),
        ('EDG-574', '[Add Connection] Step 1 - select connector type and category', 'High', 'New', 'Unassigned'),
        ('EDG-575', '[Add Connection] Step 2 - configure credentials and schedule', 'High', 'New', 'Unassigned'),
        ('EDG-576', '[Add Connection] Step 3 - set ingestion scope and filters', 'High', 'New', 'Unassigned'),
        ('EDG-577', '[Add Connection] Step 4 - test connection before saving', 'High', 'New', 'Unassigned'),
        ('EDG-578', '[Ingestion Profile] Overview tab - ingestion progress bar, run history, and quick actions', 'High', 'New', 'Unassigned'),
        ('EDG-579', '[Ingestion Profile] Configuration tab - view and edit connection settings', 'High', 'New', 'Unassigned'),
        ('EDG-580', '[Ingestion Profile] Assets tab - view indexed assets and navigate to catalog', 'High', 'New', 'Unassigned'),
        ('EDG-599', '[Add Connection] Snowflake - configure connector-specific details', 'High', 'New', 'Unassigned'),
        ('EDG-600', '[Add Connection] Databricks - configure connector-specific details', 'High', 'New', 'Unassigned'),
        ('EDG-601', '[Add Connection] Oracle - configure connector-specific details', 'High', 'New', 'Unassigned'),
        ('EDG-602', '[Add Connection] MySQL - configure connector-specific details', 'High', 'New', 'Unassigned'),
        ('EDG-603', '[Add Connection] PostgreSQL - configure connector-specific details', 'High', 'New', 'Unassigned'),
        ('EDG-604', '[Add Connection] S3 - configure connector-specific details', 'High', 'New', 'Unassigned'),
        ('EDG-605', '[Add Connection] Azure Blob - configure connector-specific details', 'High', 'New', 'Unassigned'),
        ('EDG-689', '[Connections Overview] Delete a connection', 'High', 'New', 'Unassigned'),
        ('EDG-690', '[Connections Overview] Connection health states and status model', 'High', 'New', 'Unassigned'),
    ]),
    ('Access Control', [
        ('EDG-583', 'View default roles, permission matrix, and role detail panel in Access Control', 'High', 'New', 'Unassigned'),
        ('EDG-586', 'Create an access policy with resource and operation rules', 'High', 'New', 'Unassigned'),
        ('EDG-593', 'LDAP - Role Mapping Tab: user provisioning, team sync, and role assignment', 'High', 'New', 'Unassigned'),
        ('EDG-669', 'Finalized RBAC role set - 4 roles with defaultRole and multi-role per user', 'High', 'New', 'Unassigned'),
    ]),
    ('Settings - SSO', [
        ('EDG-591', 'LDAP - Configure connection settings (Connection Tab)', 'High', 'New', 'Unassigned'),
        ('EDG-592', 'LDAP - Test Connection button and result banner', 'High', 'New', 'Unassigned'),
    ]),
    ('Connectors - Metadata Ingestion', [
        ('EDG-275', 'Snowflake Connector - Ingest Table, View & Column Metadata (MVP)', 'High', 'New', 'Unassigned'),
        ('EDG-279', 'MySQL Connector - Ingest Table, View & Column Metadata (MVP)', 'High', 'New', 'Unassigned'),
        ('EDG-281', 'Oracle Connector - Ingest Table, View & Column Metadata (MVP)', 'High', 'New', 'Unassigned'),
        ('EDG-283', 'PostgreSQL Connector - Ingest Table, View & Column Metadata (MVP)', 'High', 'New', 'Unassigned'),
        ('EDG-285', 'Azure Blob Storage Connector - Ingest File & Column Metadata (MVP)', 'High', 'New', 'Unassigned'),
    ]),
]

row = 3
stripe = False
for section, stories in sections:
    sc = SECTION_COLORS.get(section, ('FFe2e8f0', 'FF334155'))
    ws.merge_cells(f'A{row}:G{row}')
    c = ws.cell(row=row, column=1, value=f'  {section}  ({len(stories)} stories)')
    c.font = Font(bold=True, size=10, color=sc[1], name='Calibri')
    c.fill = PatternFill('solid', fgColor=sc[0])
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = thin_border()
    ws.row_dimensions[row].height = 22
    row += 1

    for key, title, priority, status, assignee in stories:
        bg = LIGHT_GRAY if stripe else WHITE
        ws.row_dimensions[row].height = 30
        cell(ws, row, 1, key, bg=bg, fg='FF3B82F6', bold=True, size=9)
        cell(ws, row, 2, section, bg=bg, fg='FF64748B', size=9)
        cell(ws, row, 3, title, bg=bg, wrap=True, size=9)
        pc = ws.cell(row=row, column=4, value=priority)
        pc.font = Font(bold=True, size=9, color=priority_colors.get(priority, 'FF64748B'), name='Calibri')
        pc.fill = PatternFill('solid', fgColor=bg)
        pc.alignment = Alignment(horizontal='center', vertical='center')
        pc.border = thin_border()
        sbg = status_bg.get(status, SLATE_SOFT)
        sfg = 'FF991B1B' if status == 'Open' else 'FF334155'
        cell(ws, row, 5, status, bg=sbg, align='center', size=9, fg=sfg, bold=(status == 'Open'))
        cell(ws, row, 6, assignee, bg=bg, size=9,
             fg='FF374151' if assignee != 'Unassigned' else 'FF94A3B8',
             italic=(assignee == 'Unassigned'))
        cell(ws, row, 7, '', bg=bg)
        stripe = not stripe
        row += 1

# ════════════════════════════════════════
# SHEET 2 — MARKED DELETE
# ════════════════════════════════════════
ws2 = wb.create_sheet('Marked DELETE')
ws2.sheet_view.showGridLines = False
ws2.freeze_panes = 'A3'
for i, w in enumerate([10, 68, 12, 30], 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.row_dimensions[1].height = 36
ws2.row_dimensions[2].height = 20

ws2.merge_cells('A1:D1')
t2 = ws2['A1']
t2.value = 'EDG  -  May 2026 Sprint  |  Marked DELETE (to be removed from backlog)'
t2.font = Font(bold=True, size=13, color=WHITE, name='Calibri')
t2.fill = PatternFill('solid', fgColor='FF7F1D1D')
t2.alignment = Alignment(horizontal='center', vertical='center')

for i, h in enumerate(['Jira ID', 'Story Title', 'Priority', 'Section'], 1):
    cell(ws2, 2, i, h, bold=True, bg='FFB91C1C', fg=WHITE, size=10, align='center')

deletes = [
    ('EDG-72',  'Asset Row Context Menu Actions in Catalog', 'Medium', 'Asset Catalog'),
    ('EDG-76',  'Certify Asset via Checklist Modal in Asset Profile', 'High', 'Asset Profile'),
    ('EDG-77',  'Post a Comment on Asset Profile', 'Medium', 'Asset Profile'),
    ('EDG-87',  'Asset Profile - View Lineage Tab', 'High', 'Asset Profile'),
    ('EDG-89',  'Asset Profile - View Activity Log Tab', 'Medium', 'Asset Profile'),
    ('EDG-91',  'Edit Asset Certification Status in Asset Profile Sidebar', 'High', 'Asset Profile'),
    ('EDG-92',  'Edit Asset Domain in Asset Profile Sidebar', 'Medium', 'Asset Profile'),
    ('EDG-93',  'Resolve a Comment on Asset Profile', 'Low', 'Asset Profile'),
    ('EDG-539', 'Asset Catalog (epic-level duplicate)', 'High', 'Epics'),
    ('EDG-540', 'Asset Profile (epic-level duplicate)', 'High', 'Epics'),
    ('EDG-547', 'Bookmark assets from catalog for My Workspace', 'Medium', 'Asset Catalog'),
    ('EDG-548', 'Bulk-apply tags to multiple assets from the catalog list', 'High', 'Asset Catalog'),
    ('EDG-549', 'Bulk update certification status for selected assets in catalog', 'Medium', 'Asset Catalog'),
    ('EDG-550', 'Export asset catalog list to CSV', 'Medium', 'Asset Catalog'),
    ('EDG-561', 'Edit or delete an existing test case (duplicate of EDG-560)', 'Medium', 'Data Quality'),
    ('EDG-572', 'Bulk apply tags to multiple assets from Tag Management', 'Medium', 'Tag Management'),
    ('EDG-581', 'Schedule automatic metadata ingestion for a connection', 'High', 'Settings - Connections'),
    ('EDG-582', 'View ingestion history and error logs for a connection', 'Medium', 'Settings - Connections'),
    ('EDG-584', 'Assign a role to a user from Access Control', 'High', 'Access Control'),
    ('EDG-585', 'Revoke a role from a user in Access Control', 'High', 'Access Control'),
    ('EDG-587', 'Edit an existing access policy', 'Medium', 'Access Control'),
    ('EDG-588', 'Delete an access policy', 'Medium', 'Access Control'),
    ('EDG-589', 'Assign an access policy to a user or team', 'High', 'Access Control'),
    ('EDG-590', 'Enforce RBAC on UI elements and API endpoints', 'High', 'Access Control'),
    ('EDG-652', 'Create and Manage Test Suites in the DQ Section', 'High', 'Data Quality'),
]

for r, (key, title, priority, section) in enumerate(deletes, 3):
    bg = RED_SOFT if r % 2 == 1 else 'FFFEF2F2'
    ws2.row_dimensions[r].height = 22
    cell(ws2, r, 1, key, bg=bg, fg='FF991B1B', bold=True, size=9)
    cell(ws2, r, 2, title, bg=bg, fg='FF7F1D1D', size=9, wrap=True)
    pc2 = ws2.cell(row=r, column=3, value=priority)
    pc2.font = Font(bold=True, size=9, color=priority_colors.get(priority, 'FF64748B'), name='Calibri')
    pc2.fill = PatternFill('solid', fgColor=bg)
    pc2.alignment = Alignment(horizontal='center', vertical='center')
    pc2.border = thin_border()
    cell(ws2, r, 4, section, bg=bg, fg='FF9CA3AF', size=9, italic=True)

# ════════════════════════════════════════
# SHEET 3 — SUMMARY
# ════════════════════════════════════════
ws3 = wb.create_sheet('Summary')
ws3.sheet_view.showGridLines = False
for i, w in enumerate([34, 14, 14, 14, 14], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.row_dimensions[1].height = 40
ws3.row_dimensions[2].height = 22

ws3.merge_cells('A1:E1')
t3 = ws3['A1']
t3.value = 'EDG May 2026 Sprint  -  Summary Dashboard'
t3.font = Font(bold=True, size=15, color=WHITE, name='Calibri')
t3.fill = PatternFill('solid', fgColor='FF0F172A')
t3.alignment = Alignment(horizontal='center', vertical='center')

for i, h in enumerate(['Section', 'Total Stories', 'High Priority', 'Assigned', 'Status'], 1):
    cell(ws3, 2, i, h, bold=True, bg='FF334155', fg=WHITE, size=10, align='center')

summary_data = [
    ('Bugs / Open Issues',              2,  2, 2, 'Open'),
    ('Settings - Teams & Users',        8,  8, 0, 'New'),
    ('Asset Catalog',                   6,  5, 0, 'New'),
    ('Asset Profile',                  10,  9, 0, 'New'),
    ('Data Quality',                   10,  7, 0, 'New'),
    ('Tag Management',                  8,  4, 0, 'New'),
    ('Settings - Connections',         17, 17, 0, 'New'),
    ('Access Control',                  4,  4, 0, 'New'),
    ('Settings - SSO',                  2,  2, 0, 'New'),
    ('Connectors - Metadata Ingestion', 5,  5, 0, 'New'),
]

stripe = False
for r, (sec, total, high, assigned, status) in enumerate(summary_data, 3):
    bg = LIGHT_GRAY if stripe else WHITE
    ws3.row_dimensions[r].height = 24
    sc = SECTION_COLORS.get(sec, ('FFe2e8f0', 'FF334155'))
    cell(ws3, r, 1, sec, bg=sc[0], fg=sc[1], bold=True, size=10)
    cell(ws3, r, 2, total, bg=bg, align='center', size=11, bold=True)
    hc = ws3.cell(row=r, column=3, value=high)
    hc.font = Font(bold=True, size=11, color='FFEF4444', name='Calibri')
    hc.fill = PatternFill('solid', fgColor=bg)
    hc.alignment = Alignment(horizontal='center', vertical='center')
    hc.border = thin_border()
    ac_val = assigned if assigned else '-'
    ac = ws3.cell(row=r, column=4, value=ac_val)
    ac.font = Font(size=10, color='FF16A34A' if assigned else 'FF94A3B8', name='Calibri', bold=bool(assigned))
    ac.fill = PatternFill('solid', fgColor=bg)
    ac.alignment = Alignment(horizontal='center', vertical='center')
    ac.border = thin_border()
    sbg = status_bg.get(status, SLATE_SOFT)
    sfg = 'FF991B1B' if status == 'Open' else 'FF334155'
    cell(ws3, r, 5, status, bg=sbg, align='center', size=9, bold=(status == 'Open'), fg=sfg)
    stripe = not stripe

total_row = len(summary_data) + 3
ws3.row_dimensions[total_row].height = 26
total_stories = sum(d[1] for d in summary_data)
total_high = sum(d[2] for d in summary_data)
total_assigned = sum(d[3] for d in summary_data)
cell(ws3, total_row, 1, 'TOTAL ACTIVE STORIES', bold=True, bg='FF1E293B', fg=WHITE, size=11)
cell(ws3, total_row, 2, total_stories, bold=True, bg='FF1E293B', fg=WHITE, size=13, align='center')
c3 = ws3.cell(row=total_row, column=3, value=total_high)
c3.font = Font(bold=True, size=13, color='FFFCA5A5', name='Calibri')
c3.fill = PatternFill('solid', fgColor='FF1E293B')
c3.alignment = Alignment(horizontal='center', vertical='center')
c3.border = thin_border()
c4 = ws3.cell(row=total_row, column=4, value=total_assigned)
c4.font = Font(bold=True, size=13, color='FF86EFAC', name='Calibri')
c4.fill = PatternFill('solid', fgColor='FF1E293B')
c4.alignment = Alignment(horizontal='center', vertical='center')
c4.border = thin_border()
cell(ws3, total_row, 5, '', bold=True, bg='FF1E293B', fg=WHITE)

note_row = total_row + 2
ws3.row_dimensions[note_row].height = 20
ws3.merge_cells(f'A{note_row}:E{note_row}')
note = ws3.cell(row=note_row, column=1, value=f'  {len(deletes)} stories marked DELETE - see "Marked DELETE" sheet')
note.font = Font(size=9, color='FF9CA3AF', italic=True, name='Calibri')
note.alignment = Alignment(horizontal='left', vertical='center')

# Reorder sheets: Summary first
wb.move_sheet('Summary', offset=-2)

path = r'C:\Users\krishna\Desktop\EDG_May2026_Sprint_Tracker.xlsx'
wb.save(path)
print('Saved:', path)
